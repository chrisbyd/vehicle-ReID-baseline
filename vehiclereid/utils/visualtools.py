from __future__ import absolute_import
from __future__ import print_function

import numpy as np
import os.path as osp
import shutil
import pandas as pd
from .iotools import mkdir_if_missing


def visualize_ranked_results(distmat, dataset, save_dir='log/ranked_results', topk=20,args = None):
    """
    Visualize ranked results
    Args:
    - distmat: distance matrix of shape (num_query, num_gallery).
    - dataset: a 2-tuple containing (query, gallery), each contains a list of (img_path, pid, camid);
               for imgreid, img_path is a string, while for vidreid, img_path is a tuple containing
               a sequence of strings.
    - save_dir: directory to save output images.
    - topk: int, denoting top-k images in the rank list to be visualized.
    """
    distmat = distmat.T
    num_q, num_g = distmat.shape

    print('Visualizing top-{} ranks'.format(topk))
    print('# query: {}\n# gallery {}'.format(num_q, num_g))
    print('Saving images to "{}"'.format(save_dir))

    query, gallery = dataset
    assert num_q == len(query)
    assert num_g == len(gallery)
    print('num of query\n',num_q)
    print('num of gallery\n',num_g)
    indices = np.argsort(distmat, axis=1)
    mkdir_if_missing(save_dir)

    def _cp_img_to(src, dst, rank, prefix):
        """
        - src: image path or tuple (for vidreid)
        - dst: target directory
        - rank: int, denoting ranked position, starting from 1
        - prefix: string
        """
        if isinstance(src, tuple) or isinstance(src, list):
            dst = osp.join(dst, prefix + '_top' + str(rank).zfill(3))
            mkdir_if_missing(dst)
            for img_path in src:
                shutil.copy(img_path, dst)
        else:
            dst = osp.join(dst, prefix + '_top' + str(rank).zfill(3) + '_name_' + osp.basename(src))
            shutil.copy(src, dst)
    save_dir = save_dir + '_' + str(args.hash_bit_number)
    for q_idx in range(20):
        qimg_path, qpid, qcamid = query[q_idx]
        if isinstance(qimg_path, tuple) or isinstance(qimg_path, list):
            qdir = osp.join(save_dir, osp.basename(qimg_path[0]))
        else:
            qdir = osp.join(save_dir, osp.basename(qimg_path))
        mkdir_if_missing(qdir)
        _cp_img_to(qimg_path, qdir, rank=0, prefix='query_{}'.format(qpid))

        rank_idx = 1
        for g_idx in indices[q_idx, :]:
            gimg_path, gpid, gcamid = gallery[g_idx]
           # invalid = (qpid == gpid) & (qcamid == gcamid)
            invalid = False
            if not invalid:
                _cp_img_to(gimg_path, qdir, rank=rank_idx, prefix='gallery_{}'.format(gpid))
                rank_idx += 1
                if rank_idx > topk:
                    break

    print("Done")






def results_to_excel(res, model_name,sheet_name):
    filename = './results.xlsx'
    df = pd.DataFrame([res], index = [model_name])
    append_df_to_excel(filename, df, sheet_name = sheet_name)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, header=False, **to_excel_kwargs)

    # save the workbook
    writer.save()
    writer.close()

if __name__ == '__main__':
    from openpyxl import load_workbook

    df = pd.DataFrame([1.2], index=['DHN'])
    filename = 'hello.xlsx'
    sheet_name = 's1'

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
    except FileNotFoundError:
        print("it doesnt exist, so we create a new file")
    df.to_excel(writer, sheet_name, startrow=0, header=False)
    writer.save()
